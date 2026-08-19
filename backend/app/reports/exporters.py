"""
Report exporters: CSV, Excel (openpyxl), PDF (reportlab) and JSON.

All four take the dictionary produced by :func:`app.reports.engine.run` and
return ``bytes``, so the API layer can stream any format without knowing the
report.

Turkish text is the hard part of exporting:

* **CSV** is written UTF-8 **with a BOM** and a semicolon delimiter, because
  that is the only combination Excel on a Turkish Windows opens correctly by
  double-click (without the BOM ``ş/ğ/İ`` arrive as mojibake).
* **PDF** needs an embedded TrueType font — the reportlab built-ins are
  Latin-1 only, so ``ş``, ``ğ``, ``İ`` render as black boxes.  We register the
  first usable font found on the machine and only fall back to ASCII folding
  when there is genuinely no TrueType font available.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from app.core.config import settings
from app.core.logging_config import get_logger
from app.core.utils import D, dumps

log = get_logger("app.reports.export")

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA_TYPE = "application/pdf"
JSON_MEDIA_TYPE = "application/json; charset=utf-8"


# ===========================================================================
# Value formatting
# ===========================================================================
def _fmt_value(value: Any, kind: str, *, for_text: bool) -> Any:
    """Coerce one cell for the target medium (text formats, numbers stay numeric)."""
    if value is None:
        return "" if for_text else None
    if kind == "date":
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y %H:%M") if for_text else value
        if isinstance(value, date):
            return value.strftime("%d.%m.%Y") if for_text else value
        return str(value)
    if kind in ("money", "number", "percent", "quantity"):
        number = D(value)
        if not for_text:
            return float(number)
        places = 3 if kind == "quantity" else 2
        text = f"{float(number):,.{places}f}".replace(",", "\u00a0")
        return f"{text}%" if kind == "percent" else text
    if kind == "integer":
        try:
            number = int(D(value))
        except (TypeError, ValueError):
            return str(value)
        return f"{number:,}".replace(",", "\u00a0") if for_text else number
    return str(value)


def _columns(result: dict[str, Any]) -> list[dict[str, Any]]:
    return list(result.get("columns") or [])


def _label(column: dict[str, Any], lang: str) -> str:
    return column.get("label") or (
        column.get("label_en") if lang == "en" else column.get("label_tr")
    ) or column.get("key", "")


def _totals_row(result: dict[str, Any], columns: list[dict[str, Any]]) -> dict[str, Any] | None:
    totals = result.get("totals") or {}
    if not totals:
        return None
    return {c["key"]: totals.get(c["key"]) for c in columns}


# ===========================================================================
# CSV
# ===========================================================================
def to_csv(result: dict[str, Any], *, lang: str = "tr", delimiter: str = ";") -> bytes:
    """
    Render as CSV, UTF-8 **with BOM**.

    The BOM is what makes Excel treat the file as UTF-8; the semicolon
    delimiter matches the Turkish list separator so columns split correctly.
    """
    columns = _columns(result)
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")

    writer.writerow([_label(c, lang) for c in columns])
    for row in result.get("rows") or []:
        writer.writerow(
            [_fmt_value(row.get(c["key"]), c.get("type", "text"), for_text=True) for c in columns]
        )

    totals = _totals_row(result, columns)
    if totals:
        first = True
        cells: list[Any] = []
        for c in columns:
            value = totals.get(c["key"])
            if value is None and first and c.get("type", "text") == "text":
                cells.append("TOPLAM" if lang != "en" else "TOTAL")
                first = False
                continue
            first = False
            cells.append(_fmt_value(value, c.get("type", "text"), for_text=True))
        writer.writerow(cells)

    return "\ufeff".encode() + buffer.getvalue().encode("utf-8")


# ===========================================================================
# Excel
# ===========================================================================
_XLSX_FORMATS = {
    "money": "#,##0.00",
    "number": "#,##0.00",
    "quantity": "#,##0.000",
    # A literal % sign: values are already scaled (12.5 means 12.5%), so the
    # built-in 0.00% format would multiply them by 100 again.
    "percent": '#,##0.00"%"',
    "integer": "#,##0",
    "date": "DD.MM.YYYY",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
_THIN = Side(style="thin", color="B4C6E7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def to_excel(result: dict[str, Any], *, title: str = "", lang: str = "tr") -> bytes:
    """Styled workbook: title block, frozen header, autofilter, totals row."""
    columns = _columns(result)
    meta = result.get("meta") or {}
    report_title = title or meta.get("title") or "Report"

    wb = Workbook()
    ws = wb.active
    ws.title = (report_title[:28] or "Rapor").replace("/", "-").replace("\\", "-")

    # --- Title block ------------------------------------------------------
    ws.cell(row=1, column=1, value=report_title).font = Font(bold=True, size=14, color="1F3864")
    subtitle_parts = [
        f"{meta.get('start', '')} — {meta.get('end', '')}" if meta.get("start") else "",
        f"{meta.get('row_count', 0)} " + ("rows" if lang == "en" else "satır"),
        ("Generated" if lang == "en" else "Oluşturma") + f": {meta.get('generated_at', '')}",
    ]
    subtitle = "   |   ".join(x for x in subtitle_parts if x)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = Font(italic=True, size=9, color="595959")
    if columns:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    header_row = 3
    for idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=_label(column, lang))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = result.get("rows") or []
    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, column in enumerate(columns, start=1):
            kind = column.get("type", "text")
            cell = ws.cell(
                row=r_idx,
                column=c_idx,
                value=_fmt_value(row.get(column["key"]), kind, for_text=False),
            )
            cell.border = _BORDER
            if kind in _XLSX_FORMATS:
                cell.number_format = _XLSX_FORMATS[kind]
            cell.alignment = Alignment(
                horizontal=column.get("align", "left"), vertical="center"
            )

    totals = _totals_row(result, columns)
    if totals:
        t_idx = header_row + len(rows) + 1
        labelled = False
        for c_idx, column in enumerate(columns, start=1):
            kind = column.get("type", "text")
            value = totals.get(column["key"])
            if value is None and not labelled and kind == "text":
                value = "TOTAL" if lang == "en" else "TOPLAM"
                labelled = True
                cell = ws.cell(row=t_idx, column=c_idx, value=value)
            else:
                cell = ws.cell(
                    row=t_idx, column=c_idx, value=_fmt_value(value, kind, for_text=False)
                )
                if kind in _XLSX_FORMATS:
                    cell.number_format = _XLSX_FORMATS[kind]
            cell.font = Font(bold=True)
            cell.fill = _TOTAL_FILL
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal=column.get("align", "left"))

    # --- Layout -----------------------------------------------------------
    for idx, column in enumerate(columns, start=1):
        longest = len(_label(column, lang))
        for row in rows[:500]:  # sampling keeps very long reports fast
            text = _fmt_value(row.get(column["key"]), column.get("type", "text"), for_text=True)
            longest = max(longest, len(str(text)))
        ws.column_dimensions[get_column_letter(idx)].width = min(
            52, max(9, min(longest + 2, int(column.get("width", 16)) + 12))
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if columns and rows:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"
        )
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


# ===========================================================================
# PDF
# ===========================================================================
#: Candidate TrueType fonts, best Unicode coverage first.
_FONT_CANDIDATES: tuple[tuple[str, str, str], ...] = (
    ("DejaVuSans", "DejaVuSans.ttf", "DejaVuSans-Bold.ttf"),
    ("Calibri", "calibri.ttf", "calibrib.ttf"),
    ("Arial", "arial.ttf", "arialbd.ttf"),
    ("Tahoma", "tahoma.ttf", "tahomabd.ttf"),
    ("Verdana", "verdana.ttf", "verdanab.ttf"),
)

_FONT_DIRS: tuple[Path, ...] = (
    Path("C:/Windows/Fonts"),
    Path("/usr/share/fonts/truetype/dejavu"),
    Path("/usr/share/fonts/truetype"),
    Path("/Library/Fonts"),
    settings.data_path / "fonts",
)

#: Resolved once per process: (regular font name, bold font name, unicode ok).
_FONTS: tuple[str, str, bool] | None = None

_TR_FOLD = str.maketrans(
    {
        "ç": "c", "Ç": "C", "ğ": "g", "Ğ": "G", "ı": "i", "İ": "I",
        "ö": "o", "Ö": "O", "ş": "s", "Ş": "S", "ü": "u", "Ü": "U",
        "\u00a0": " ", "—": "-", "–": "-", "…": "...",
    }
)


def _register_fonts() -> tuple[str, str, bool]:
    """
    Find and register a Unicode-capable TrueType font.

    Returns ``(regular, bold, unicode_ok)``.  ``unicode_ok=False`` means the
    caller must ASCII-fold Turkish characters before drawing them.
    """
    global _FONTS
    if _FONTS is not None:
        return _FONTS

    for family, regular_file, bold_file in _FONT_CANDIDATES:
        for directory in _FONT_DIRS:
            regular_path = directory / regular_file
            if not regular_path.is_file():
                continue
            bold_path = directory / bold_file
            try:
                pdfmetrics.registerFont(TTFont(family, str(regular_path)))
                bold_name = family
                if bold_path.is_file():
                    bold_name = f"{family}-Bold"
                    pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
                pdfmetrics.registerFontFamily(
                    family, normal=family, bold=bold_name, italic=family, boldItalic=bold_name
                )
            except Exception:  # pragma: no cover - depends on the host's fonts
                log.warning("Could not register font %s from %s", family, regular_path)
                continue
            _FONTS = (family, bold_name, True)
            return _FONTS

    log.warning("No TrueType font found — PDF export falls back to Helvetica with folded text")
    _FONTS = ("Helvetica", "Helvetica-Bold", False)
    return _FONTS


def _pdf_text(value: str, unicode_ok: bool) -> str:
    text = str(value)
    if not unicode_ok:
        text = text.translate(_TR_FOLD)
    return (
        text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )


def to_pdf(
    result: dict[str, Any],
    *,
    title: str = "",
    subtitle: str = "",
    lang: str = "tr",
    company: str | None = None,
) -> bytes:
    """Landscape A4 PDF: company header, striped table, totals row, page numbers."""
    regular, bold, unicode_ok = _register_fonts()
    columns = _columns(result)
    meta = result.get("meta") or {}
    rows = result.get("rows") or []

    report_title = title or meta.get("title") or "Report"
    company_name = company or settings.app_name
    generated = meta.get("generated_at") or datetime.now().isoformat(timespec="seconds")
    if not subtitle and meta.get("start"):
        subtitle = f"{meta.get('start')} — {meta.get('end')}"

    page_size = landscape(A4)
    page_width, page_height = page_size
    margin = 12 * mm
    stream = io.BytesIO()

    doc = BaseDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin + 12 * mm,
        bottomMargin=margin + 8 * mm,
        title=report_title,
        author=company_name,
    )

    header_company = _pdf_text(company_name, unicode_ok)
    header_title = _pdf_text(report_title, unicode_ok)
    header_generated = _pdf_text(
        ("Generated" if lang == "en" else "Oluşturma") + f": {generated}", unicode_ok
    )
    page_label = "Page" if lang == "en" else "Sayfa"

    def _decorate(canvas: Any, _doc: Any) -> None:
        canvas.saveState()
        canvas.setFont(bold, 10)
        canvas.setFillColor(colors.HexColor("#1F3864"))
        canvas.drawString(margin, page_height - margin - 2 * mm, header_company)
        canvas.setFont(regular, 8)
        canvas.setFillColor(colors.HexColor("#595959"))
        canvas.drawRightString(
            page_width - margin, page_height - margin - 2 * mm, header_generated
        )
        canvas.setStrokeColor(colors.HexColor("#1F3864"))
        canvas.setLineWidth(0.8)
        canvas.line(
            margin, page_height - margin - 4 * mm, page_width - margin, page_height - margin - 4 * mm
        )
        canvas.setFont(regular, 8)
        canvas.drawCentredString(
            page_width / 2, margin - 1 * mm, f"{page_label} {canvas.getPageNumber()}"
        )
        canvas.drawString(margin, margin - 1 * mm, header_title)
        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        leftPadding=0,
        rightPadding=0,
        topPadding=0,
        bottomPadding=0,
        id="body",
    )
    doc.addPageTemplates([PageTemplate(id="report", frames=[frame], onPage=_decorate)])

    title_style = ParagraphStyle(
        "ReportTitle", fontName=bold, fontSize=14, leading=17,
        textColor=colors.HexColor("#1F3864"), spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", fontName=regular, fontSize=9, leading=12,
        textColor=colors.HexColor("#595959"), spaceAfter=6,
    )
    cell_style = ParagraphStyle("Cell", fontName=regular, fontSize=7, leading=8.6)
    head_style = ParagraphStyle(
        "Head", fontName=bold, fontSize=7, leading=8.6, textColor=colors.white
    )

    story: list[Any] = [
        Paragraph(_pdf_text(report_title, unicode_ok), title_style),
    ]
    info = " | ".join(
        x
        for x in (
            subtitle,
            f"{meta.get('row_count', len(rows))} " + ("rows" if lang == "en" else "satır"),
        )
        if x
    )
    if info:
        story.append(Paragraph(_pdf_text(info, unicode_ok), subtitle_style))
    story.append(Spacer(1, 2 * mm))

    if not columns:
        doc.build(story)
        return stream.getvalue()

    header_cells = [Paragraph(_pdf_text(_label(c, lang), unicode_ok), head_style) for c in columns]
    data: list[list[Any]] = [header_cells]
    for row in rows:
        line: list[Any] = []
        for column in columns:
            text = _fmt_value(row.get(column["key"]), column.get("type", "text"), for_text=True)
            line.append(Paragraph(_pdf_text(text, unicode_ok), cell_style))
        data.append(line)

    totals = _totals_row(result, columns)
    total_row_index: int | None = None
    if totals:
        line = []
        labelled = False
        for column in columns:
            value = totals.get(column["key"])
            kind = column.get("type", "text")
            if value is None and not labelled and kind == "text":
                text = "TOTAL" if lang == "en" else "TOPLAM"
                labelled = True
            else:
                labelled = True
                text = _fmt_value(value, kind, for_text=True)
            line.append(
                Paragraph(f"<b>{_pdf_text(text, unicode_ok)}</b>", cell_style)
            )
        data.append(line)
        total_row_index = len(data) - 1

    # Proportional widths from the declared column widths, scaled to the frame.
    declared = [max(6, int(c.get("width", 16))) for c in columns]
    scale = doc.width / sum(declared)
    widths = [w * scale for w in declared]

    style_commands: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B4C6E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for idx, column in enumerate(columns):
        align = column.get("align", "left").upper()
        style_commands.append(("ALIGN", (idx, 0), (idx, -1), align if align in ("LEFT", "RIGHT", "CENTER") else "LEFT"))
    if total_row_index is not None:
        style_commands.append(
            ("BACKGROUND", (0, total_row_index), (-1, total_row_index), colors.HexColor("#D9E2F3"))
        )
        style_commands.append(
            ("LINEABOVE", (0, total_row_index), (-1, total_row_index), 0.8, colors.HexColor("#1F3864"))
        )

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle(style_commands))
    story.append(table)

    doc.build(story)
    return stream.getvalue()


# ===========================================================================
# JSON
# ===========================================================================
def to_json(result: dict[str, Any]) -> bytes:
    """Raw result as UTF-8 JSON (Decimals become numbers, dates ISO strings)."""
    return dumps(result, indent=2).encode("utf-8")


# ===========================================================================
# Dispatch
# ===========================================================================
FORMATS: dict[str, str] = {
    "csv": CSV_MEDIA_TYPE,
    "excel": XLSX_MEDIA_TYPE,
    "xlsx": XLSX_MEDIA_TYPE,
    "pdf": PDF_MEDIA_TYPE,
    "json": JSON_MEDIA_TYPE,
}

_EXTENSIONS = {"csv": "csv", "excel": "xlsx", "xlsx": "xlsx", "pdf": "pdf", "json": "json"}


def extension_for(fmt: str) -> str:
    return _EXTENSIONS.get(fmt.lower(), "txt")


def render(
    result: dict[str, Any],
    fmt: str,
    *,
    title: str = "",
    subtitle: str = "",
    lang: str = "tr",
    company: str | None = None,
) -> tuple[bytes, str]:
    """Render *result* in *fmt*, returning ``(payload, media_type)``."""
    kind = (fmt or "csv").lower()
    if kind == "csv":
        return to_csv(result, lang=lang), CSV_MEDIA_TYPE
    if kind in ("excel", "xlsx"):
        return to_excel(result, title=title, lang=lang), XLSX_MEDIA_TYPE
    if kind == "pdf":
        return (
            to_pdf(result, title=title, subtitle=subtitle, lang=lang, company=company),
            PDF_MEDIA_TYPE,
        )
    if kind == "json":
        return to_json(result), JSON_MEDIA_TYPE
    raise ValueError(f"unsupported export format: {fmt}")


__all__ = [
    "FORMATS",
    "extension_for",
    "render",
    "to_csv",
    "to_excel",
    "to_json",
    "to_pdf",
]
